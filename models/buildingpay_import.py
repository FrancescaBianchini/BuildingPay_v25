# -*- coding: utf-8 -*-
import base64
import logging
from io import BytesIO
from datetime import date, timedelta
from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

IMPORT_STATES = [
    ('draft', 'Bozza'),
    ('done', 'Elaborato'),
    ('error', 'Errori'),
]


class BuildingPayImport(models.Model):
    """
    Gestisce l'importazione di un file Excel per la generazione
    di fatture di vendita e ordini di acquisto (retrocessioni).

    Il file Excel deve contenere le colonne:
    - id_esterno_condominio
    - data_fattura
    - quantita
    - descrizione_riga
    - prezzo_unitario
    """
    _name = 'buildingpay_v25.import'
    _description = 'Importazione Fatture Condomini'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'create_date desc'
    _rec_name = 'name'

    # -------------------------------------------------------
    # Campi base
    # -------------------------------------------------------
    name = fields.Char(
        string='Riferimento importazione',
        required=True,
        copy=False,
        default=lambda self: _('Nuova importazione'),
        readonly=True,
        # Odoo 17: states={} rimosso - readonly gestito nella vista con invisible/readonly
    )
    state = fields.Selection(
        selection=IMPORT_STATES,
        string='Stato',
        default='draft',
        required=True,
        tracking=True,
        readonly=True,
    )
    excel_file = fields.Binary(
        string='File Excel',
        required=True,
        attachment=True,
        # Odoo 17: states={} rimosso - readonly gestito nella vista
    )
    excel_filename = fields.Char(
        string='Nome file',
        # Odoo 17: states={} rimosso - readonly gestito nella vista
    )
    user_id = fields.Many2one(
        comodel_name='res.users',
        string='Caricato da',
        default=lambda self: self.env.user,
        readonly=True,
    )
    upload_date = fields.Datetime(
        string='Data caricamento',
        default=fields.Datetime.now,
        readonly=True,
    )
    note = fields.Text(
        string='Note',
    )

    # -------------------------------------------------------
    # Risultati elaborazione
    # -------------------------------------------------------
    error_ids = fields.One2many(
        comodel_name='buildingpay_v25.import.error',
        inverse_name='import_id',
        string='Errori',
        readonly=True,
    )
    error_count = fields.Integer(
        string='Numero errori',
        compute='_compute_error_count',
    )
    invoice_ids = fields.Many2many(
        comodel_name='account.move',
        string='Fatture generate',
        readonly=True,
    )
    invoice_count = fields.Integer(
        string='Fatture generate',
        compute='_compute_invoice_count',
    )
    purchase_order_ids = fields.Many2many(
        comodel_name='purchase.order',
        string='Ordini di acquisto generati',
        readonly=True,
    )
    purchase_order_count = fields.Integer(
        string='Ordini di acquisto',
        compute='_compute_purchase_order_count',
    )

    # -------------------------------------------------------
    # Compute
    # -------------------------------------------------------
    @api.depends('error_ids')
    def _compute_error_count(self):
        for rec in self:
            rec.error_count = len(rec.error_ids)

    @api.depends('invoice_ids')
    def _compute_invoice_count(self):
        for rec in self:
            rec.invoice_count = len(rec.invoice_ids)

    @api.depends('purchase_order_ids')
    def _compute_purchase_order_count(self):
        for rec in self:
            rec.purchase_order_count = len(rec.purchase_order_ids)

    # -------------------------------------------------------
    # Sequenza
    # -------------------------------------------------------
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', _('Nuova importazione')) == _('Nuova importazione'):
                vals['name'] = self.env['ir.sequence'].next_by_code(
                    'buildingpay_v25.import') or _('Nuova importazione')
        return super().create(vals_list)

    # -------------------------------------------------------
    # Azioni
    # -------------------------------------------------------
    def action_process(self):
        """
        Avvia l'elaborazione del file Excel.
        Genera fatture di vendita e ordini di acquisto.
        """
        self.ensure_one()
        if self.state != 'draft':
            raise UserError(_('Solo i record in stato "Bozza" possono essere elaborati.'))
        if not self.excel_file:
            raise UserError(_('Caricare un file Excel prima di procedere.'))

        # Pulisci errori precedenti
        self.error_ids.unlink()

        errors = []
        invoices_created = self.env['account.move']
        purchase_orders_created = self.env['purchase.order']

        try:
            import openpyxl
            file_data = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(BytesIO(file_data))
            ws = wb.active

            # Leggi intestazioni (prima riga)
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[str(cell.value).strip().lower()] = col_idx

            required_cols = [
                'id_esterno_condominio', 'data_fattura',
                'quantita', 'descrizione_riga', 'prezzo_unitario',
            ]
            missing = [c for c in required_cols if c not in headers]
            if missing:
                raise UserError(_(
                    'Colonne mancanti nel file Excel: %s'
                ) % ', '.join(missing))

            # Raggruppa righe per condominio+data (ogni combinazione = una fattura)
            invoice_groups = {}  # key: (condominio_id, data_fattura)

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):
                    continue  # Salta righe vuote

                def get_val(col_name):
                    idx = headers.get(col_name)
                    if idx is None:
                        return None
                    return row[idx - 1] if idx <= len(row) else None

                ext_id = get_val('id_esterno_condominio')
                data_fattura_raw = get_val('data_fattura')
                quantita = get_val('quantita')
                descrizione = get_val('descrizione_riga')
                prezzo_unitario = get_val('prezzo_unitario')

                # Validazioni riga
                row_errors = []

                # 1. Trova condominio tramite external ID
                condominio = None
                if not ext_id:
                    row_errors.append((row_idx, 'E001', _('ID esterno condominio mancante.')))
                else:
                    ir_data = self.env['ir.model.data'].sudo().search([
                        ('complete_name', '=', str(ext_id)),
                        ('model', '=', 'res.partner'),
                    ], limit=1)
                    if not ir_data:
                        row_errors.append((row_idx, 'E002', _(
                            'Condominio con ID esterno "%s" non trovato.'
                        ) % ext_id))
                    else:
                        condominio = self.env['res.partner'].browse(ir_data.res_id)
                        if not condominio.exists():
                            row_errors.append((row_idx, 'E003', _(
                                'Condominio con ID %s non esiste.'
                            ) % ir_data.res_id))
                        elif condominio.type != 'condominio':
                            row_errors.append((row_idx, 'E004', _(
                                'Il partner "%s" non è di tipo "Condominio".'
                            ) % condominio.name))

                # 2. Valida data fattura
                invoice_date = None
                if not data_fattura_raw:
                    row_errors.append((row_idx, 'E005', _('Data fattura mancante.')))
                else:
                    try:
                        if isinstance(data_fattura_raw, (date,)):
                            invoice_date = data_fattura_raw
                        else:
                            from datetime import datetime
                            invoice_date = datetime.strptime(
                                str(data_fattura_raw), '%Y-%m-%d').date()
                    except Exception:
                        row_errors.append((row_idx, 'E006', _(
                            'Formato data non valido: %s. Usare YYYY-MM-DD.'
                        ) % data_fattura_raw))

                # 3. Valida quantità
                try:
                    qty = float(quantita or 0)
                    if qty <= 0:
                        raise ValueError()
                except (TypeError, ValueError):
                    row_errors.append((row_idx, 'E007', _(
                        'Quantità non valida: %s.'
                    ) % quantita))
                    qty = 0.0

                if row_errors:
                    errors.extend(row_errors)
                    continue

                # Raggruppa per condominio + data
                key = (condominio.id, invoice_date)
                if key not in invoice_groups:
                    invoice_groups[key] = {
                        'condominio': condominio,
                        'invoice_date': invoice_date,
                        'lines': [],
                    }
                invoice_groups[key]['lines'].append({
                    'descrizione': str(descrizione or ''),
                    'quantita': qty,
                    'prezzo_unitario': float(prezzo_unitario or 0.0),
                    'row_idx': row_idx,
                })

            # ----------------------------------------
            # Genera fatture e ordini di acquisto
            # ----------------------------------------
            # Raggruppa fatture per amministratore (per creare un PO per admin)
            admin_invoices = {}  # key: admin.id → lista fatture

            for key, group_data in invoice_groups.items():
                condominio = group_data['condominio']
                admin = condominio.parent_id
                if not admin or not admin.is_amministratore:
                    errors.append((0, 'E010', _(
                        'Il condominio "%s" non ha un amministratore padre.'
                    ) % condominio.name))
                    continue

                # Trova prodotto PagoPa
                pagopa_product = self.env['product.template'].search([
                    ('is_condominio_pagopa', '=', True),
                ], limit=1)
                if not pagopa_product:
                    errors.append((0, 'E011', _(
                        'Nessun prodotto con flag "Condominio PagoPa" attivo trovato.'
                    )))
                    continue

                product_product = pagopa_product.product_variant_id

                # Ottieni prezzo dal listino dell'amministratore
                pricelist = admin.property_product_pricelist
                lines_vals = []
                for line in group_data['lines']:
                    if pricelist:
                        price = pricelist.get_product_price(
                            product=product_product,
                            quantity=line['quantita'],
                            partner=condominio,
                            date=group_data['invoice_date'],
                        )
                    else:
                        price = line['prezzo_unitario']

                    lines_vals.append((0, 0, {
                        'product_id': product_product.id,
                        'name': line['descrizione'] or pagopa_product.name,
                        'quantity': line['quantita'],
                        'price_unit': price,
                        'tax_ids': [(6, 0, product_product.taxes_id.ids)],
                    }))

                # Crea fattura di vendita
                invoice = self.env['account.move'].create({
                    'move_type': 'out_invoice',
                    'partner_id': condominio.id,
                    'invoice_date': group_data['invoice_date'],
                    'invoice_line_ids': lines_vals,
                })
                invoices_created |= invoice

                # Raggruppa per amministratore
                if admin.id not in admin_invoices:
                    admin_invoices[admin.id] = {'admin': admin, 'invoices': []}
                admin_invoices[admin.id]['invoices'].append(invoice)

            # ----------------------------------------
            # Crea ordini di acquisto per amministratori
            # ----------------------------------------
            admin_po_product = self._get_or_create_retrocessione_product()

            for admin_id, data in admin_invoices.items():
                admin = data['admin']
                pricelist = admin.property_product_pricelist
                perc_admin = pricelist.perc_retrocessione_amministratore if pricelist else 0.0

                if perc_admin > 0:
                    po_lines_vals = []
                    for inv in data['invoices']:
                        condominio_name = inv.partner_id.name
                        imponibile = inv.amount_untaxed
                        retro_price = (perc_admin / 100.0) * imponibile
                        po_lines_vals.append((0, 0, {
                            'product_id': admin_po_product.id,
                            'name': _(
                                'Retrocessione riconosciuta all\'amministratore '
                                'per la fattura del condominio %s'
                            ) % condominio_name,
                            'product_qty': 1.0,
                            'price_unit': retro_price,
                            'date_planned': fields.Datetime.now(),
                            'product_uom': admin_po_product.uom_po_id.id,
                        }))
                    if po_lines_vals:
                        po = self.env['purchase.order'].create({
                            'partner_id': admin.id,
                            'order_line': po_lines_vals,
                        })
                        purchase_orders_created |= po

                # ----------------------------------------
                # Ordine di acquisto per il referrer
                # ----------------------------------------
                referrer = admin.referrer_id
                if referrer:
                    perc_referrer = pricelist.perc_retrocessione_referrer if pricelist else 0.0
                    if perc_referrer > 0:
                        po_ref_lines = []
                        for inv in data['invoices']:
                            condominio_name = inv.partner_id.name
                            imponibile = inv.amount_untaxed
                            retro_price = (perc_referrer / 100.0) * imponibile
                            po_ref_lines.append((0, 0, {
                                'product_id': admin_po_product.id,
                                'name': _(
                                    'Retrocessione riconosciuta al referrer '
                                    'per la fattura del condominio %s'
                                ) % condominio_name,
                                'product_qty': 1.0,
                                'price_unit': retro_price,
                                'date_planned': fields.Datetime.now(),
                                'product_uom': admin_po_product.uom_po_id.id,
                            }))
                        if po_ref_lines:
                            po_ref = self.env['purchase.order'].create({
                                'partner_id': referrer.id,
                                'order_line': po_ref_lines,
                            })
                            purchase_orders_created |= po_ref

        except UserError:
            raise
        except Exception as e:
            _logger.error('BuildingPay Import: errore generico: %s', e)
            errors.append((0, 'E999', _('Errore generico: %s') % str(e)))

        # ----------------------------------------
        # Salva risultati
        # ----------------------------------------
        error_vals = []
        for err in errors:
            row_num, code, desc = err
            error_vals.append((0, 0, {
                'row_number': row_num,
                'error_code': code,
                'error_description': desc,
            }))

        new_state = 'error' if errors else 'done'
        write_vals = {
            'state': new_state,
            'error_ids': error_vals,
            'invoice_ids': [(6, 0, invoices_created.ids)],
            'purchase_order_ids': [(6, 0, purchase_orders_created.ids)],
        }
        self.write(write_vals)

        # Messaggio di riepilogo
        if errors:
            self.message_post(body=_(
                '<b>Elaborazione completata con %d errori.</b><br/>'
                'Fatture create: %d | Ordini acquisto creati: %d'
            ) % (len(errors), len(invoices_created), len(purchase_orders_created)))
        else:
            self.message_post(body=_(
                '<b>Elaborazione completata con successo.</b><br/>'
                'Fatture create: %d | Ordini acquisto creati: %d'
            ) % (len(invoices_created), len(purchase_orders_created)))

    def _get_or_create_retrocessione_product(self):
        """
        Restituisce il prodotto da usare per le righe degli ordini di acquisto
        di retrocessione. Lo crea se non esiste.
        """
        product = self.env['product.product'].search([
            ('default_code', '=', 'BUILDINGPAY_RETRO'),
        ], limit=1)
        if not product:
            product_template = self.env['product.template'].create({
                'name': _('Retrocessione BuildingPay'),
                'default_code': 'BUILDINGPAY_RETRO',
                'type': 'service',
                'purchase_ok': True,
                'sale_ok': False,
            })
            product = product_template.product_variant_id
        return product

    def action_reset_draft(self):
        """Riporta in bozza per una nuova elaborazione."""
        self.ensure_one()
        self.write({
            'state': 'draft',
            'error_ids': [(5, 0, 0)],
            'invoice_ids': [(5, 0, 0)],
            'purchase_order_ids': [(5, 0, 0)],
        })

    def action_view_invoices(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Fatture generate'),
            'res_model': 'account.move',
            'view_mode': 'list,form',
            'domain': [('id', 'in', self.invoice_ids.ids)],
        }

    def action_view_purchase_orders(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Ordini di acquisto'),
            'res_model': 'purchase.order',
            'view_mode': 'list,form',
            'domain': [('id', 'in', self.purchase_order_ids.ids)],
        }
