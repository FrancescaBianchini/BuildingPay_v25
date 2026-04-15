# -*- coding: utf-8 -*-
import uuid
import logging
from datetime import date, timedelta
from odoo import api, fields, models, _
from odoo.exceptions import ValidationError, UserError

_logger = logging.getLogger(__name__)


class ResPartner(models.Model):
    """
    Estensione di res.partner per BuildingPay v15.

    Aggiunge:
    - Tipo contatto 'Condominio' (con icona edificio nel form backend)
    - Flag Amministratore BuildingPay
    - Codice Referrer (8 car.) + link di invito computato
    - Dati bancari: IBAN e banca tramite res.partner.bank nativo Odoo
    - Listino: usa property_product_pricelist nativo Odoo
    - Sezione portale 'Contratti' con due documenti:
        1. Accordo retrocessioni amministratore ED
           (placeholder: nome, CF, IBAN, banca, data)
        2. Accordo Condomini Aggregati ED
           (placeholder: nome, CF + Allegato A auto-compilato con condomini attivi)
    - Flag: accordo_retrocessioni_ed, accordo_condomini_aggregati_ed
    - Campi italiani: fiscalcode, pec_mail, codice_destinatario
    """
    _inherit = 'res.partner'

    # -------------------------------------------------------
    # Tipo contatto esteso con 'condominio'
    # -------------------------------------------------------
    type = fields.Selection(
        selection_add=[('condominio', 'Condominio')],
        ondelete={'condominio': 'set default'},
    )

    # -------------------------------------------------------
    # Flag e dati Amministratore
    # -------------------------------------------------------
    is_amministratore = fields.Boolean(
        string='Amministratore',
        default=False,
        # tracking rimosso intenzionalmente: con tracking=True Odoo tenta di creare
        # un mail.message al write(). Per i nuovi utenti portale appena creati
        # questa operazione fallisce e causa un rollback silenzioso dell'intero
        # write() via savepoint, impedendo il salvataggio di tutti i campi.
    )
    referrer_code = fields.Char(
        string='Codice Referrer',
        copy=False,
        index=True,
        size=8,
        help=(
            'Codice univoco di 8 caratteri per il link di invito. '
            'Viene auto-generato quando si abilita il flag Amministratore. '
            'Può essere modificato manualmente o rigenerato con il pulsante.'
        ),
    )
    referral_url = fields.Char(
        string='Link di invito',
        compute='_compute_referral_url',
        store=False,
        help="URL completo che l'amministratore può inviare ai suoi clienti per registrarsi.",
    )
    referrer_id = fields.Many2one(
        comodel_name='res.partner',
        string='Referrer',
        ondelete='set null',
        tracking=True,
    )
    referred_ids = fields.One2many(
        comodel_name='res.partner',
        inverse_name='referrer_id',
        string='Amministratori invitati',
    )

    # -------------------------------------------------------
    # Privacy
    # -------------------------------------------------------
    privacy_accepted = fields.Boolean(
        string='Privacy accettata',
        default=False,
        # tracking rimosso intenzionalmente: stessa ragione di is_amministratore.
        help='True se il contatto ha accettato la privacy policy al momento della registrazione.',
    )
    privacy_accepted_date = fields.Datetime(
        string='Data accettazione privacy',
        readonly=True,
    )

    # -------------------------------------------------------
    # CONTRATTO 1: Accordo retrocessioni amministratore ED
    # Placeholder nel template:
    #   [NOME AMMINISTRATORE]  → partner.name
    #   [CODICE FISCALE]       → partner.fiscalcode
    #   [IBAN]                 → primo res.partner.bank.acc_number
    #   [NOME BANCA]           → bank_id.name su res.partner.bank
    #   [DATA]                 → data odierna DD/MM/YYYY
    # -------------------------------------------------------
    accordo_retrocessioni_ed = fields.Boolean(
        string='Accordo retrocessioni amministratore ED',
        default=False,
        tracking=True,
        help='Attivato quando l\'amministratore carica l\'Accordo Retrocessioni firmato.',
    )
    accordo_retrocessioni_file = fields.Binary(
        string='File Accordo Retrocessioni',
        attachment=True,
    )
    accordo_retrocessioni_filename = fields.Char(
        string='Nome file accordo retrocessioni',
    )
    accordo_retrocessioni_upload_date = fields.Datetime(
        string='Data caricamento accordo retrocessioni',
        readonly=True,
    )

    # -------------------------------------------------------
    # CONTRATTO 2: Accordo Condomini Aggregati ED
    # Placeholder nel template:
    #   [NOME AMMINISTRATORE]  → partner.name
    #   [________]             → partner.fiscalcode
    #   [ALLEGATO_A]           → tabella condomini attivi (nome | indirizzo | IBAN)
    # -------------------------------------------------------
    accordo_condomini_aggregati_ed = fields.Boolean(
        string='Accordo condomini aggregati ED',
        default=False,
        tracking=True,
        help='Attivato quando l\'amministratore carica l\'Accordo Condomini Aggregati firmato.',
    )
    accordo_condomini_file = fields.Binary(
        string='File Accordo Condomini Aggregati',
        attachment=True,
    )
    accordo_condomini_filename = fields.Char(
        string='Nome file accordo condomini',
    )
    accordo_condomini_upload_date = fields.Datetime(
        string='Data caricamento accordo condomini',
        readonly=True,
    )

    # -------------------------------------------------------
    # Data archiviazione (per indirizzi di tipo condominio)
    # -------------------------------------------------------
    data_archiviazione = fields.Date(
        string='Data archiviazione',
        readonly=True,
    )

    # -------------------------------------------------------
    # Campi italiani (compatibili con l10n_it_edi)
    # -------------------------------------------------------
    fiscalcode = fields.Char(
        string='Codice Fiscale',
        size=16,
        index=True,
    )
    pec_mail = fields.Char(
        string='Email PEC',
    )
    codice_destinatario = fields.Char(
        string='Codice Destinatario SDI',
        size=7,
    )
    electronic_invoice_subjected = fields.Boolean(
        string='Soggetto a fatturazione elettronica',
        default=False,
    )
    electronic_invoice_obliged_subject = fields.Boolean(
        string='Obbligo fatturazione elettronica',
        default=False,
    )

    # -------------------------------------------------------
    # Condomini figli (indirizzi di tipo condominio)
    # -------------------------------------------------------
    condominio_ids = fields.One2many(
        comodel_name='res.partner',
        inverse_name='parent_id',
        string='Condomini',
        domain=[('type', '=', 'condominio'), ('active', '=', True)],
    )
    condominio_count = fields.Integer(
        string='Numero Condomini',
        compute='_compute_condominio_count',
    )

    # -------------------------------------------------------
    # Compute
    # -------------------------------------------------------
    @api.depends('child_ids', 'child_ids.type', 'child_ids.active')
    def _compute_condominio_count(self):
        for partner in self:
            partner.condominio_count = self.env['res.partner'].search_count([
                ('parent_id', '=', partner.id),
                ('type', '=', 'condominio'),
                ('active', '=', True),
            ])

    # -------------------------------------------------------
    # Create / Write
    # -------------------------------------------------------
    @api.model_create_multi
    def create(self, vals_list):
        # Se is_amministratore=True e il chiamante non ha già passato un referrer_code,
        # lo generiamo qui direttamente nel vals per evitare un secondo write().
        for vals in vals_list:
            if vals.get('is_amministratore') and not vals.get('referrer_code'):
                vals['referrer_code'] = self._generate_referrer_code()
        return super().create(vals_list)

    def write(self, vals):
        result = super().write(vals)
        # Genera referrer_code per gli amministratori che non ce l'hanno ancora.
        # Usiamo tracking_disable=True per evitare che questa scrittura secondaria
        # tenti di creare mail.message (che potrebbe fallire per nuovi utenti portale).
        needs_code = self.filtered(
            lambda p: p.is_amministratore and not p.referrer_code)
        if needs_code:
            for partner in needs_code:
                partner.with_context(
                    tracking_disable=True, mail_notrack=True
                ).write({'referrer_code': self._generate_referrer_code()})
        return result

    @api.depends('referrer_code')
    def _compute_referral_url(self):
        """Calcola il link di invito: {base_url}/web/signup?referrer={code}.
        Non condizionato da is_amministratore: il link e' sempre disponibile
        se esiste un codice referrer, indipendentemente dal flag.
        """
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url', '')
        base_url = base_url.rstrip('/')
        for partner in self:
            if partner.referrer_code:
                partner.referral_url = '%s/web/signup?referrer=%s' % (
                    base_url, partner.referrer_code)
            else:
                partner.referral_url = False

    def action_generate_referrer_code(self):
        """Genera un nuovo codice referrer casuale di 8 caratteri.
        Ritorna False in modo che Odoo 17 ricarichi il form e mostri subito
        il nuovo codice senza dover fare refresh manuale della pagina.
        """
        self.ensure_one()
        self.referrer_code = self._generate_referrer_code()
        # False → Odoo 17 aggiorna automaticamente il record nel form corrente
        return False

    @api.model
    def _generate_referrer_code(self):
        """Genera un codice alfanumerico casuale di 8 caratteri (maiuscolo)."""
        return uuid.uuid4().hex[:8].upper()

    # -------------------------------------------------------
    # Upload: Accordo Retrocessioni
    # -------------------------------------------------------
    def action_upload_retrocessioni(self, file_data, filename):
        """
        Chiamato dal portale quando l'utente carica l'Accordo Retrocessioni firmato.
        """
        self.ensure_one()
        self.write({
            'accordo_retrocessioni_file': file_data,
            'accordo_retrocessioni_filename': filename,
            'accordo_retrocessioni_ed': True,
            'accordo_retrocessioni_upload_date': fields.Datetime.now(),
        })

    # -------------------------------------------------------
    # Upload: Accordo Condomini Aggregati
    # -------------------------------------------------------
    def action_upload_accordo_condomini(self, file_data, filename):
        """
        Chiamato dal portale quando l'utente carica l'Accordo Condomini Aggregati firmato.
        Attiva il flag accordo_condomini_aggregati_ed e crea eventuale attività automatica.
        """
        self.ensure_one()
        self.write({
            'accordo_condomini_file': file_data,
            'accordo_condomini_filename': filename,
            'accordo_condomini_aggregati_ed': True,
            'accordo_condomini_upload_date': fields.Datetime.now(),
        })
        self._create_contratto_activity()

    def _create_contratto_activity(self):
        """Crea attività automatica se configurata."""
        self.ensure_one()
        config = self.env['buildingpay_v25.config'].get_config_for_website()
        if not config or not config.create_activity_on_contract:
            return
        if not config.activity_responsible_id:
            return

        deadline = date.today() + timedelta(days=config.activity_days or 5)
        activity_type = self.env.ref(
            'mail.mail_activity_data_todo', raise_if_not_found=False)
        self.activity_schedule(
            activity_type_id=activity_type.id if activity_type else False,
            summary=_('Controllare il contratto Accordo Condomini Aggregati '
                       'caricato dall\'amministratore'),
            date_deadline=deadline,
            user_id=config.activity_responsible_id.id,
        )

    # -------------------------------------------------------
    # Archiviazione condominio
    # -------------------------------------------------------
    def action_archive_condominio(self):
        """Archivia un indirizzo di tipo condominio."""
        self.ensure_one()
        if self.type != 'condominio':
            raise UserError(_(
                'Solo gli indirizzi di tipo "Condominio" possono essere archiviati.'))
        self.write({
            'data_archiviazione': fields.Date.today(),
            'active': False,
        })
        self._send_condominio_dismesso_email()

    def _send_condominio_dismesso_email(self):
        """Genera Excel condominio dismesso e lo invia via email."""
        self.ensure_one()
        config = self.env['buildingpay_v25.config'].get_config_for_website()
        if not config or not config.condomini_dismessi_email:
            return

        try:
            import openpyxl
            from io import BytesIO
            import base64

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Condominio Dismesso'
            headers = [
                'ID Esterno Amministratore', 'Nome Amministratore',
                'ID Esterno Condominio', 'Nome Condominio',
                'Indirizzo Completo', 'Dismesso',
            ]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h).font = \
                    openpyxl.styles.Font(bold=True)

            admin = self.parent_id
            admin_ext = self.env['ir.model.data'].search([
                ('model', '=', 'res.partner'),
                ('res_id', '=', admin.id if admin else 0),
            ], limit=1)
            cond_ext = self.env['ir.model.data'].search([
                ('model', '=', 'res.partner'), ('res_id', '=', self.id),
            ], limit=1)
            address = ' '.join(filter(None, [
                self.street, self.zip, self.city,
                self.state_id.name if self.state_id else '',
                self.country_id.name if self.country_id else '',
            ]))
            row = [
                admin_ext.complete_name if admin_ext else '',
                admin.name if admin else '',
                cond_ext.complete_name if cond_ext else '',
                self.name or '', address, True,
            ]
            for col, v in enumerate(row, 1):
                ws.cell(row=2, column=col, value=v)

            out = BytesIO()
            wb.save(out)
            recipients = [
                e.strip() for e in config.condomini_dismessi_email.split(',')
                if e.strip()
            ]
            if recipients:
                self.env['mail.mail'].sudo().create({
                    'subject': _('Condominio dismesso: %s') % self.name,
                    'body_html': _(
                        '<p>Il condominio <b>%s</b> è stato dismesso '
                        'in data %s.</p>') % (self.name, fields.Date.today()),
                    'email_to': ','.join(recipients),
                    'attachment_ids': [(0, 0, {
                        'name': 'condominio_dismesso_%s.xlsx' % self.name,
                        'datas': base64.b64encode(out.getvalue()),
                        'mimetype': ('application/vnd.openxmlformats-officedocument'
                                     '.spreadsheetml.sheet'),
                    })],
                }).send()
        except Exception as e:
            _logger.error('BuildingPay v21: errore email condominio dismesso: %s', e)

    # -------------------------------------------------------
    # Azione pianificata: report giornaliero condomini attivi
    # -------------------------------------------------------
    @api.model
    def action_send_daily_condomini_report(self):
        """Genera e invia il report Excel giornaliero dei condomini attivi."""
        try:
            import openpyxl
            from io import BytesIO
            import base64

            condominii = self.search([
                ('type', '=', 'condominio'),
                ('active', '=', True),
                ('parent_id', '!=', False),
                ('parent_id.is_amministratore', '=', True),
            ])
            if not condominii:
                _logger.info('BuildingPay v21: nessun condominio attivo.')
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Condomini Attivi'
            headers = [
                'ID Esterno Amministratore', 'Nome Amministratore',
                'ID Esterno Condominio', 'Nome Condominio', 'Indirizzo',
                'IBAN', 'Email PEC', 'Codice Destinatario', 'Codice Fiscale',
            ]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h).font = \
                    openpyxl.styles.Font(bold=True)

            for row_idx, condo in enumerate(condominii, 2):
                admin = condo.parent_id
                admin_ext = self.env['ir.model.data'].search([
                    ('model', '=', 'res.partner'), ('res_id', '=', admin.id),
                ], limit=1)
                cond_ext = self.env['ir.model.data'].search([
                    ('model', '=', 'res.partner'), ('res_id', '=', condo.id),
                ], limit=1)
                bank = self.env['res.partner.bank'].search([
                    ('partner_id', '=', condo.id),
                ], limit=1)
                address = ' '.join(filter(None, [
                    condo.street, condo.zip, condo.city,
                    condo.state_id.name if condo.state_id else '',
                ]))
                row_data = [
                    admin_ext.complete_name if admin_ext else '',
                    admin.name or '',
                    cond_ext.complete_name if cond_ext else '',
                    condo.name or '', address,
                    bank.acc_number if bank else '',
                    condo.pec_mail or '',
                    condo.codice_destinatario or '',
                    condo.fiscalcode or '',
                ]
                for col, v in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col, value=v)

            out = BytesIO()
            wb.save(out)
            excel_data = base64.b64encode(out.getvalue())

            configs = self.env['buildingpay_v25.config'].search([
                ('condomini_attivati_email', '!=', False),
            ])
            all_recipients = set()
            for cfg in configs:
                for email in cfg.condomini_attivati_email.split(','):
                    if email.strip():
                        all_recipients.add(email.strip())

            if all_recipients:
                today_str = fields.Date.today().strftime('%Y-%m-%d')
                self.env['mail.mail'].sudo().create({
                    'subject': _('Report Condomini Attivi - %s') % today_str,
                    'body_html': _(
                        '<p>Report giornaliero condomini attivi del %s.</p>'
                    ) % today_str,
                    'email_to': ','.join(all_recipients),
                    'attachment_ids': [(0, 0, {
                        'name': 'condomini_attivi_%s.xlsx' % today_str,
                        'datas': excel_data,
                        'mimetype': ('application/vnd.openxmlformats-officedocument'
                                     '.spreadsheetml.sheet'),
                    })],
                }).send()
        except Exception as e:
            _logger.error('BuildingPay v21: errore report giornaliero: %s', e)
